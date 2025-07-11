import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
import openpyxl
from io import BytesIO

# Set page configuration
st.set_page_config(
    page_title="Airline Bids Analysis",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for professional styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.8rem;
        font-weight: 700;
        color: #1e3a8a;
        text-align: center;
        margin-bottom: 1.5rem;
        text-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .section-header {
        font-size: 1.6rem;
        font-weight: 600;
        color: #1f2937;
        margin: 2rem 0 1rem 0;
        border-bottom: 3px solid #3b82f6;
        padding-bottom: 0.5rem;
    }
    .executive-summary {
        background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
        padding: 2rem;
        border-radius: 12px;
        border-left: 6px solid #3b82f6;
        margin: 1.5rem 0;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
    }
    .insight-box {
        background-color: #f0f9ff;
        padding: 1.5rem;
        border-radius: 8px;
        border-left: 4px solid #0284c7;
        margin: 1rem 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    .metric-card {
        background: white;
        padding: 1.5rem;
        border-radius: 8px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border: 1px solid #e5e7eb;
    }
    .rating-legend {
        background: #fafafa;
        padding: 1rem;
        border-radius: 8px;
        border: 1px solid #d1d5db;
        margin: 1rem 0;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 24px;
        background-color: #f8fafc;
        padding: 8px;
        border-radius: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        padding: 0 24px;
        border-radius: 6px;
        font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)

@st.cache_data
def load_data(uploaded_file):
    """Load and process the Excel file data"""
    try:
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        
        if 'Airline Bids' not in workbook.sheetnames:
            st.error("Sheet 'Airline Bids' not found in the Excel file")
            return None
        
        sheet = workbook['Airline Bids']
        
        # Convert to DataFrame starting from row 11 (data starts there)
        data = []
        headers = []
        
        # Get headers from row 10
        for col in range(3, sheet.max_column + 1):
            cell = sheet.cell(row=11, column=col)
            headers.append(cell.value if cell.value else f'col_{col}')
        
        # Get data starting from row 12
        for row in range(12, sheet.max_row + 1):
            row_data = []
            for col in range(3, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                row_data.append(cell.value)
            
            # Only add rows that have data in key columns
            if row_data[3] and row_data[4] and row_data[13]:  # Origin Airport, Destination Airport, Airline
                data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Clean and standardize column names
        column_mapping = {
            'Commodity Group': 'commodity_group',
            'TempControlled': 'temp_controlled',
            'Air Mode': 'air_mode',
            'Origin Airport': 'origin_airport',
            'Destination Airport': 'destination_airport',
            'Origin Country': 'origin_country',
            'Destinatin Country': 'destination_country',
            'Origin Region': 'origin_region',
            'Destination Region': 'destination_region',
            'Airline': 'airline',
            'Intention to Bid (Yes/No)': 'intention_to_bid',
            'Direct / Indirect': 'direct_indirect',
            'Via': 'via',
            'Currency': 'currency',
            'Min Charge': 'min_charge',
            'Min Charge2': 'min_charge2',
            'Percentage': 'percentage',
            'Numerical Rating': 'rating',
            'Column1': 'rating_category'  # This contains Green/Orange/Red
        }
        
        # Rename columns that exist in the DataFrame
        for old_name, new_name in column_mapping.items():
            if old_name in df.columns:
                df = df.rename(columns={old_name: new_name})
        
        # Convert numeric columns
        numeric_columns = ['min_charge2', 'rating']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Create route column
        df['route'] = df['origin_airport'] + ' → ' + df['destination_airport']
        
        # Create color mapping based on BOTH numerical rating AND rating_category
        def get_color_from_data(row):
            # First try to use rating_category (Green/Red/Orange)
            if pd.notna(row.get('rating_category')) and str(row.get('rating_category')).strip() != 'None':
                category = str(row.get('rating_category')).strip().lower()
                if category == 'green':
                    return '#22c55e'  # Bright Green
                elif category == 'orange':
                    return '#f97316'  # Bright Orange  
                elif category == 'red':
                    return '#ef4444'  # Bright Red
            
            # If rating_category is not available, use numerical rating
            rating = row.get('rating')
            if pd.notna(rating):
                if rating == 1:
                    return '#22c55e'  # Bright Green
                elif rating == 2:
                    return '#f97316'  # Bright Orange
                elif rating == 3:
                    return '#ef4444'  # Bright Red
            
            return '#6b7280'  # Gray for unknown
        
        df['color'] = df.apply(get_color_from_data, axis=1)
        
        # Clean rating category
        if 'rating_category' in df.columns:
            df['rating_category'] = df['rating_category'].astype(str).str.strip()
            df['rating_category'] = df['rating_category'].replace({'nan': 'Unknown', '': 'Unknown'})
        
        # Filter out rows with missing critical data
        df = df.dropna(subset=['origin_airport', 'destination_airport', 'airline', 'min_charge2'])
        
        return df
        
    except Exception as e:
        st.error(f"Error loading data: {str(e)}")
        return None

def show_executive_overview(df):
    """Show executive summary of the data"""
    st.markdown('<h1 class="main-header">✈️ Airline Bids Analysis Dashboard</h1>', unsafe_allow_html=True)
    
    # Executive Summary
    st.markdown("""
    <div class="executive-summary">
    <h3>📊 Executive Summary</h3>
    <p><strong>Purpose:</strong> This dashboard analyzes competitive airline pricing bids for cargo shipments across global routes to identify cost optimization opportunities and carrier performance.</p>
    
    <p><strong>Key Benefits:</strong></p>
    <ul>
        <li>🎯 <strong>Cost Optimization:</strong> Identify the most competitive pricing options for each route</li>
        <li>📈 <strong>Carrier Performance:</strong> Evaluate airline competitiveness and service coverage</li>
        <li>💡 <strong>Strategic Insights:</strong> Make data-driven decisions for logistics partnerships</li>
        <li>⚡ <strong>Quick Analysis:</strong> Instantly compare options for specific origin-destination pairs</li>
    </ul>
    </div>
    """, unsafe_allow_html=True)
    
    # Only show metrics if data is available
    if not df.empty and 'route' in df.columns:
        # Key Metrics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_routes = df['route'].nunique()
            st.markdown(f"""
            <div class="metric-card">
            <h4 style="color: #1e40af; margin: 0;">🌍 Global Routes</h4>
            <h2 style="color: #1f2937; margin: 0.5rem 0;">{total_routes:,}</h2>
            <p style="color: #6b7280; margin: 0;">Origin-destination pairs</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            total_airlines = df['airline'].nunique()
            st.markdown(f"""
            <div class="metric-card">
            <h4 style="color: #1e40af; margin: 0;">✈️ Carrier Network</h4>
            <h2 style="color: #1f2937; margin: 0.5rem 0;">{total_airlines:,}</h2>
            <p style="color: #6b7280; margin: 0;">Competing airlines</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            avg_price = df['min_charge2'].mean()
            st.markdown(f"""
            <div class="metric-card">
            <h4 style="color: #1e40af; margin: 0;">💰 Average Rate</h4>
            <h2 style="color: #1f2937; margin: 0.5rem 0;">${avg_price:.2f}</h2>
            <p style="color: #6b7280; margin: 0;">Per shipment</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            if 'rating' in df.columns:
                best_rate_pct = (df['rating'] == 1).mean() * 100
            else:
                best_rate_pct = 0
            st.markdown(f"""
            <div class="metric-card">
            <h4 style="color: #1e40af; margin: 0;">🎯 Optimization Rate</h4>
            <h2 style="color: #1f2937; margin: 0.5rem 0;">{best_rate_pct:.1f}%</h2>
            <p style="color: #6b7280; margin: 0;">Best pricing options</p>
            </div>
            """, unsafe_allow_html=True)

def create_route_analysis(df, origin, destination):
    """Create detailed analysis for a specific route"""
    route_data = df[(df['origin_airport'] == origin) & (df['destination_airport'] == destination)].copy()
    
    if route_data.empty:
        st.warning("⚠️ No carriers serve this route in our current bid data.")
        return None
    
    route_name = f"{origin} → {destination}"
    
    # Route Performance Metrics
    st.markdown(f'<div class="section-header">📍 Route Analysis: {route_name}</div>', unsafe_allow_html=True)
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        carrier_count = route_data['airline'].nunique()
        st.metric("🏢 Carriers Available", carrier_count)
    
    with col2:
        best_price = route_data['min_charge2'].min()
        st.metric("💰 Best Rate", f"${best_price:.2f}")
    
    with col3:
        avg_price = route_data['min_charge2'].mean()
        st.metric("📊 Market Average", f"${avg_price:.2f}")
    
    with col4:
        if len(route_data) > 1:
            price_variance = route_data['min_charge2'].max() - route_data['min_charge2'].min()
            st.metric("📈 Price Spread", f"${price_variance:.2f}")
        else:
            st.metric("📈 Price Spread", "$0.00")
    
    # Carrier Comparison Chart
    st.markdown("### 🏆 Carrier Competitiveness Analysis")
    
    # Sort by price for better visualization
    route_data = route_data.sort_values('min_charge2')
    
    # Force color assignment based on what we see in the data
    def assign_colors_manually(row):
        # Use rating_category if available
        if pd.notna(row['rating_category']) and str(row['rating_category']).strip().lower() in ['green', 'red', 'orange']:
            category = str(row['rating_category']).strip().lower()
            if category == 'green':
                return '#22c55e'
            elif category == 'orange': 
                return '#f97316'
            elif category == 'red':
                return '#ef4444'
        return '#6b7280'
    
    route_data['display_color'] = route_data.apply(assign_colors_manually, axis=1)
    
    # Create professional chart with FORCED COLORS
    fig = go.Figure()
    
    # Add bars with explicit color list
    colors_list = route_data['display_color'].tolist()
    
    fig.add_trace(go.Bar(
        x=route_data['airline'],
        y=route_data['min_charge2'],
        marker=dict(
            color=colors_list,  # Use explicit color list
            line=dict(width=1, color='rgba(0,0,0,0.1)')
        ),
        text=[f"${price:.2f}" for price in route_data['min_charge2']],
        textposition='outside',
        textfont=dict(size=12, color='#1f2937'),
        hovertemplate="<b>%{x}</b><br>" +
                      "Rate: $%{y:.2f}<br>" +
                      "Rating: %{customdata}<br>" +
                      "<extra></extra>",
        customdata=route_data['rating'],
        name="Shipping Rate"
    ))
    
    fig.update_layout(
        title=dict(
            text=f"Carrier Pricing Comparison - {route_name}",
            font=dict(size=16, color='#1f2937'),
            x=0.5
        ),
        xaxis_title="Airlines",
        yaxis_title="Rate (USD)",
        height=450,
        showlegend=False,
        plot_bgcolor='white',
        paper_bgcolor='white',
        xaxis=dict(
            categoryorder='total ascending',
            gridcolor='#f3f4f6',
            title_font=dict(size=14, color='#374151')
        ),
        yaxis=dict(
            gridcolor='#f3f4f6',
            title_font=dict(size=14, color='#374151')
        ),
        margin=dict(t=60, b=60, l=60, r=60)
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Rating System Explanation
    st.markdown("""
    <div class="rating-legend">
    <h4 style="margin-top: 0;">🎨 Competitive Rating System</h4>
    <div style="display: flex; gap: 2rem; flex-wrap: wrap;">
        <div><span style="color: #10b981; font-size: 1.2rem;">●</span> <strong>Rating 1:</strong> Most Competitive (Best Value)</div>
        <div><span style="color: #f59e0b; font-size: 1.2rem;">●</span> <strong>Rating 2:</strong> Market Rate (Fair Value)</div>
        <div><span style="color: #ef4444; font-size: 1.2rem;">●</span> <strong>Rating 3:</strong> Premium Pricing (Highest Cost)</div>
    </div>
    </div>
    """, unsafe_allow_html=True)
    
    return route_data

def show_carrier_insights(route_data, route_name):
    """Show detailed carrier insights and recommendations"""
    
    if len(route_data) == 0:
        return
    
    # Strategic Insights
    st.markdown("### 💡 Strategic Insights & Recommendations")
    
    best_option = route_data.loc[route_data['min_charge2'].idxmin()]
    worst_option = route_data.loc[route_data['min_charge2'].idxmax()] if len(route_data) > 1 else best_option
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown(f"""
        <div class="insight-box">
        <h4>🏆 Recommended Carrier</h4>
        <p><strong>Airline:</strong> {best_option['airline']}</p>
        <p><strong>Rate:</strong> ${best_option['min_charge2']:.2f}</p>
        <p><strong>Service:</strong> {best_option.get('direct_indirect', 'N/A')}</p>
        <p><strong>Competitiveness:</strong> Rating {int(best_option['rating']) if pd.notna(best_option['rating']) else 'N/A'}</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        if len(route_data) > 1:
            savings = worst_option['min_charge2'] - best_option['min_charge2']
            savings_pct = (savings / worst_option['min_charge2']) * 100
            
            st.markdown(f"""
            <div class="insight-box">
            <h4>💰 Cost Optimization</h4>
            <p><strong>Potential Savings:</strong> ${savings:.2f}</p>
            <p><strong>Savings Percentage:</strong> {savings_pct:.1f}%</p>
            <p><strong>vs. Highest Bidder:</strong> {worst_option['airline']}</p>
            <p><strong>Annual Impact:</strong> ${savings * 52:.0f} (weekly shipments)</p>
            </div>
            """, unsafe_allow_html=True)
    
    # Detailed Carrier Information
    st.markdown("### 📋 Detailed Carrier Comparison")
    
    # Prepare display data
    display_columns = ['airline', 'min_charge2', 'rating', 'rating_category', 'direct_indirect']
    available_columns = [col for col in display_columns if col in route_data.columns]
    
    display_df = route_data[available_columns].copy()
    
    # Rename for professional display
    column_names = {
        'airline': 'Carrier',
        'min_charge2': 'Rate (USD)',
        'rating': 'Competitive Rating',
        'rating_category': 'Market Position',
        'direct_indirect': 'Service Type'
    }
    
    display_df = display_df.rename(columns=column_names)
    
    # Format pricing
    if 'Rate (USD)' in display_df.columns:
        display_df['Rate (USD)'] = display_df['Rate (USD)'].apply(lambda x: f"${x:.2f}")
    
    # Style the dataframe
    def highlight_ratings(row):
        if 'Competitive Rating' in row.index and pd.notna(row['Competitive Rating']):
            rating = row['Competitive Rating']
            if rating == 1:
                return ['background-color: #d1fae5; border-left: 4px solid #10b981'] * len(row)
            elif rating == 2:
                return ['background-color: #fef3c7; border-left: 4px solid #f59e0b'] * len(row)
            elif rating == 3:
                return ['background-color: #fee2e2; border-left: 4px solid #ef4444'] * len(row)
        return ['background-color: #f9fafb'] * len(row)
    
    styled_df = display_df.style.apply(highlight_ratings, axis=1)
    st.dataframe(styled_df, use_container_width=True, hide_index=True)

def create_airlines_overview(df):
    """Create comprehensive airlines performance overview"""
    st.markdown('<div class="section-header">🏢 Carrier Performance Dashboard</div>', unsafe_allow_html=True)
    
    # Calculate airline statistics step by step to avoid confusion
    avg_rates = df.groupby('airline')['min_charge2'].mean().round(2)
    routes_covered = df.groupby('airline')['route'].nunique()
    total_bids = df.groupby('airline').size()  # Count rows per airline
    
    # Combine into one dataframe
    airline_stats = pd.DataFrame({
        'airline': avg_rates.index,
        'avg_rate': avg_rates.values,
        'routes_covered': routes_covered.values,
        'total_bids': total_bids.values
    })
    
    # Sort by total bids
    airline_stats = airline_stats.sort_values('total_bids', ascending=False)
    
    # Performance Summary Table
    st.markdown("### 📊 Carrier Performance Summary")
    
    # Format for executive presentation
    display_stats = airline_stats.copy()
    display_stats['avg_rate'] = display_stats['avg_rate'].apply(lambda x: f"${x:.2f}")
    
    # Professional column names
    display_stats = display_stats.rename(columns={
        'airline': 'Carrier',
        'routes_covered': 'Routes Covered',
        'total_bids': 'Total Bids',
        'avg_rate': 'Average Rate'
    })
    
    # Show the 4 columns in correct order
    display_stats = display_stats[['Carrier', 'Routes Covered', 'Total Bids', 'Average Rate']]
    
    st.dataframe(display_stats, use_container_width=True, hide_index=True)
    
    # Performance Analysis Charts
    col1, col2 = st.columns(2)
    
    with col1:
        # Market Coverage vs Pricing
        fig1 = px.scatter(
            airline_stats.head(15),
            x='routes_covered',
            y='avg_rate',
            size='total_bids',
            hover_name='airline',
            title="Market Coverage vs Average Pricing",
            labels={
                'routes_covered': 'Routes Covered',
                'avg_rate': 'Average Rate (USD)',
                'total_bids': 'Total Bids'
            }
        )
        fig1.update_layout(
            height=400,
            plot_bgcolor='white',
            paper_bgcolor='white'
        )
        st.plotly_chart(fig1, use_container_width=True)
    
    with col2:
        # Top carriers by total bids
        top_carriers = airline_stats.nlargest(10, 'total_bids')
        
        fig2 = px.bar(
            top_carriers,
            x='airline',
            y='total_bids',
            title='Most Active Carriers (Total Bids)',
            labels={'total_bids': 'Total Bids', 'airline': 'Carriers'}
        )
        fig2.update_layout(
            height=400,
            showlegend=False,
            plot_bgcolor='white',
            paper_bgcolor='white'
        )
        st.plotly_chart(fig2, use_container_width=True)
    
    # Market Insights
    st.markdown("""
    <div class="insight-box">
    <h4>📈 Market Analysis Insights</h4>
    <p><strong>Chart Interpretation:</strong></p>
    <ul>
        <li><strong>Left Chart:</strong> Shows carrier market coverage (x-axis) vs pricing levels (y-axis). Bubble size indicates bid volume.</li>
        <li><strong>Right Chart:</strong> Ranks carriers by their total number of bids submitted.</li>
        <li><strong>Strategic Value:</strong> Identify carriers that offer both broad coverage and competitive pricing for partnership opportunities.</li>
    </ul>
    </div>
    """, unsafe_allow_html=True)

def main():
    # File upload
    uploaded_file = st.file_uploader(
        "📁 Upload Airline Bids Excel File",
        type=['xlsx', 'xls'],
        help="Select the Excel file containing the 'Airline Bids' sheet"
    )
    
    if uploaded_file is not None:
        # Load data
        with st.spinner("🔄 Processing bid data..."):
            df = load_data(uploaded_file)
        
        if df is not None:
            # Show executive overview
            show_executive_overview(df)
            
            # Create main navigation tabs
            tab1, tab2 = st.tabs(["🎯 Route Analysis", "🏢 Carrier Performance"])
            
            with tab1:
                st.markdown('<div class="section-header">🛫 Route-Specific Analysis</div>', unsafe_allow_html=True)
                
                st.markdown("""
                <div class="insight-box">
                <h4>📍 Route Selection</h4>
                <p>Select origin and destination airports to analyze carrier options, pricing competitiveness, and identify cost optimization opportunities for specific shipping lanes.</p>
                </div>
                """, unsafe_allow_html=True)
                
                # Airport selection
                origins = sorted(df['origin_airport'].unique())
                
                col1, col2 = st.columns(2)
                
                with col1:
                    selected_origin = st.selectbox(
                        "🛫 Origin Airport",
                        origins,
                        help="Select departure airport"
                    )
                
                with col2:
                    # Filter destinations based on origin
                    available_destinations = df[df['origin_airport'] == selected_origin]['destination_airport'].unique()
                    available_destinations = sorted(available_destinations)
                    
                    selected_destination = st.selectbox(
                        "🛬 Destination Airport",
                        available_destinations,
                        help="Select arrival airport"
                    )
                
                # Route analysis
                if selected_origin and selected_destination:
                    route_data = create_route_analysis(df, selected_origin, selected_destination)
                    
                    if route_data is not None and not route_data.empty:
                        # Analysis tabs
                        sub_tab1, sub_tab2 = st.tabs(["📊 Overview", "📋 Detailed Analysis"])
                        
                        with sub_tab1:
                            route_name = f"{selected_origin} → {selected_destination}"
                            show_carrier_insights(route_data, route_name)
                        
                        with sub_tab2:
                            st.markdown("### 🔍 Comprehensive Route Data")
                            st.dataframe(route_data, use_container_width=True)
                            
                            # Download option
                            csv = route_data.to_csv(index=False)
                            st.download_button(
                                label="📥 Download Route Analysis",
                                data=csv,
                                file_name=f"route_analysis_{selected_origin}_{selected_destination}.csv",
                                mime="text/csv"
                            )
            
            with tab2:
                create_airlines_overview(df)
    
    else:
        # Professional landing page - create empty dataframe
        empty_df = pd.DataFrame()
        show_executive_overview(empty_df)
        
        st.markdown("""
        <div class="insight-box">
        <h4>🚀 Getting Started</h4>
        <p><strong>Upload your airline bids Excel file to begin analysis.</strong></p>
        <p>This executive dashboard provides comprehensive insights into:</p>
        <ul>
            <li>✅ <strong>Cost Optimization:</strong> Identify best pricing options across routes</li>
            <li>✅ <strong>Carrier Performance:</strong> Evaluate airline competitiveness and coverage</li>
            <li>✅ <strong>Strategic Analysis:</strong> Data-driven recommendations for logistics partnerships</li>
            <li>✅ <strong>Market Intelligence:</strong> Competitive landscape overview and trends</li>
        </ul>
        </div>
        """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
